from flask import Flask, request, jsonify, send_from_directory, send_file
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import pandas as pd
import os, io, json

app = Flask(__name__, static_folder="static")
UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

state = {"fixably": None, "actual": None, "stockedout": None}
FILE_KINDS = ("fixably", "actual", "stockedout")

NORM = lambda s: str(s).strip().upper().replace(" ", "")

def read_file(path):
    if path.endswith(".xlsx"):
        return pd.read_excel(path, dtype=str).fillna("")
    return pd.read_csv(path, dtype=str).fillna("")

def norm(series):
    return series.str.replace(r'\s+', '', regex=True).str.upper()

@app.route("/")
def index():
    return send_from_directory("static", "index.html")

@app.route("/upload", methods=["POST"])
def upload():
    kind = request.form["kind"]  # "fixably", "actual", "stockedout"
    f = request.files["file"]
    path = os.path.join(UPLOAD_DIR, f"{kind}_{f.filename}")
    f.save(path)
    df = read_file(path)
    state[kind] = {"path": path, "name": f.filename, "columns": list(df.columns), "df": df}
    return jsonify({"columns": list(df.columns), "preview": df.head(3).to_dict(orient="records")})

@app.route("/state")
def get_state():
    _rehydrate()
    out = {}
    for kind in FILE_KINDS:
        v = state.get(kind)
        if v:
            out[kind] = {"name": v["name"], "columns": v["columns"],
                         "preview": v["df"].head(3).to_dict(orient="records")}
    if state.get("last_reconcile"):
        out["last_reconcile"] = state["last_reconcile"]
    return jsonify(out)

def _rehydrate():
    """Reload the most recent file per kind from disk into memory (survives restarts)."""
    for kind in FILE_KINDS:
        if state.get(kind):
            continue
        files = [f for f in os.listdir(UPLOAD_DIR) if f.startswith(f"{kind}_")]
        if not files:
            continue
        latest = max(files, key=lambda f: os.path.getmtime(os.path.join(UPLOAD_DIR, f)))
        path = os.path.join(UPLOAD_DIR, latest)
        try:
            df = read_file(path)
        except Exception:
            continue
        name = latest
        while name.startswith(f"{kind}_"):
            name = name[len(kind) + 1:]
        state[kind] = {"path": path, "name": name,
                       "columns": list(df.columns), "df": df}

@app.route("/reset", methods=["POST"])
def reset():
    for kind in FILE_KINDS:
        state[kind] = None
        for f in os.listdir(UPLOAD_DIR):
            if f.startswith(f"{kind}_"):
                try:
                    os.remove(os.path.join(UPLOAD_DIR, f))
                except OSError:
                    pass
    state["last_reconcile"] = None
    return jsonify({"ok": True})

@app.route("/outtake", methods=["POST"])
def outtake():
    if not state.get("stockedout"):
        return jsonify({"error": "Upload your Stocked Out sheet first."}), 400
    df = state["stockedout"]["df"]
    cols = {c.strip().upper(): c for c in df.columns}
    body = request.json or {}
    m = body.get("map", {})
    only = body.get("serials")
    if only is None:
        return jsonify({"error": "Run the comparison first — FOR OUTTAKE uses its results."}), 400
    only = set(only)
    ser_c = m.get("serial") or cols.get("SERIAL NUMBER", "")
    part_c = m.get("part") or cols.get("PART NUMBER", "")
    desc_c = m.get("description") or cols.get("DESCRIPTION", "")
    ref_c = m.get("reference") or cols.get("REFERENCE NUMBER", "") or cols.get("AR", "")

    wb = Workbook()
    ws = wb.active
    ws.title = "FOR OUTTAKE"
    ws.append(["FOR OUTTAKE"])
    ws.append(["PART NUMBER", "DESCRIPTION", "REFERENCE NUMBER", "REMARKS"])
    n = 0
    for _, r in df.iterrows():
        serial = str(r[ser_c]).strip() if ser_c else ""
        if not serial or NORM(serial) not in only:
            continue
        ws.append([r[part_c] if part_c else "", r[desc_c] if desc_c else "",
                   r[ref_c] if ref_c else "", serial])
        n += 1

    # Non-serialized Fixably items not in the actual on-hand sheet
    nonser = (state.get("last_reconcile") or {}).get("nonser_outtake", [])
    if nonser:
        ws.append([])
        ws.append(["NON-SERIALIZED — NOT IN ACTUAL ON HAND"])
        ws.append(["PART NUMBER", "DESCRIPTION", "QUANTITY"])
        for it in nonser:
            ws.append([it["code"], it["description"], it["quantity"]])

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    resp = send_file(buf, as_attachment=True, download_name="FOR_OUTTAKE.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp.headers["X-Stats"] = json.dumps({"outtake": n, "nonser": len(nonser)})
    return resp

@app.route("/reconcile", methods=["POST"])
def reconcile():
    body = request.json
    fx_map = body["fixably_map"]
    ac_map = body["actual_map"]

    if not state.get("fixably") or not state.get("actual"):
        return jsonify({"error": "Upload both the Fixably export and the Actual / On Hand sheet first."}), 400

    fx_df, ac_df = state["fixably"]["df"], state["actual"]["df"]
    for label, df, mp in [("Fixably", fx_df, fx_map), ("Actual / On Hand", ac_df, ac_map)]:
        for field in ("serial", "code", "quantity"):
            col = mp.get(field)
            if not col or col not in df.columns:
                return jsonify({"error": f"{label}: column for '{field}' not found. Re-check the mapping."}), 400

    fx = fx_df.rename(columns={
        fx_map["serial"]: "serial", fx_map["code"]: "code", fx_map["quantity"]: "quantity"
    })
    ac = ac_df.rename(columns={
        ac_map["serial"]: "serial", ac_map["code"]: "code", ac_map["quantity"]: "quantity"
    })

    fx["serial"] = norm(fx["serial"])
    ac["serial"] = norm(ac["serial"])

    fx_serials = set(fx[fx["serial"] != ""]["serial"])
    ac_serials = set(ac[ac["serial"] != ""]["serial"])

    # Stocked-out serials: in Fixably but already used — exclude from discrepancy
    so_serials = set()
    so_ref = {}
    if state["stockedout"] and body.get("stockedout_map"):
        so_raw = state["stockedout"]["df"]
        so_cols = {c.strip().upper(): c for c in so_raw.columns}
        ref_c = so_cols.get("REFERENCE NUMBER", "") or so_cols.get("AR", "")
        so_df = so_raw.rename(columns={body["stockedout_map"]["serial"]: "serial"})
        so_df["serial"] = norm(so_df["serial"])
        so_serials = set(so_df[so_df["serial"] != ""]["serial"])
        if ref_c:
            so_ref = {s: str(r).strip() for s, r in zip(so_df["serial"], so_df[ref_c]) if s}

    in_actual_not_fixably = sorted(ac_serials - fx_serials)
    in_fixably_not_actual = sorted((fx_serials - ac_serials) - so_serials)
    stocked_out_excluded  = sorted((fx_serials - ac_serials) & so_serials)

    # Non-serialized Fixably items (no serial) whose code is NOT in the actual on-hand sheet
    fx_name_col = next((c for c in fx_df.columns if c.strip().lower() in ("name", "description")), None)
    ac_codes = set(str(c).strip() for c in ac["code"] if str(c).strip())
    nonser_outtake = {}
    for _, r in fx_df.iterrows():
        if str(r[fx_map["serial"]]).strip():
            continue
        code = str(r[fx_map["code"]]).strip()
        if not code or code in ac_codes:
            continue
        item = nonser_outtake.setdefault(code, {"code": code,
            "description": str(r[fx_name_col]).strip() if fx_name_col else "", "quantity": 0})
        q = str(r[fx_map["quantity"]]).strip()
        item["quantity"] += int(q) if q.lstrip("-").isdigit() else 0
    nonser_outtake = list(nonser_outtake.values())

    # Quantity mismatch by part code
    def qty_sum(df):
        return df.groupby("code")["quantity"].apply(
            lambda x: sum(int(v) for v in x if str(v).isdigit())
        ).reset_index()

    merged = pd.merge(
        qty_sum(fx).rename(columns={"quantity": "fx_qty"}),
        qty_sum(ac).rename(columns={"quantity": "ac_qty"}),
        on="code", how="outer"
    ).fillna(0)
    merged["fx_qty"] = merged["fx_qty"].astype(int)
    merged["ac_qty"] = merged["ac_qty"].astype(int)
    qty_mismatch = merged[merged["fx_qty"] != merged["ac_qty"]].to_dict(orient="records")

    state["serial_lookup"] = dict(zip(fx["serial"], fx["code"]))

    result = {
        "in_actual_not_fixably": in_actual_not_fixably,
        "in_fixably_not_actual": in_fixably_not_actual,
        "stocked_out_excluded": stocked_out_excluded,
        "stocked_out_ref": so_ref,
        "nonser_outtake": nonser_outtake,
        "qty_mismatch": qty_mismatch,
        "summary": {
            "fx_total": len(fx_serials),
            "ac_total": len(ac_serials),
            "matched": len(fx_serials & ac_serials),
            "_debug": {
                "fx_col": fx_map["serial"], "fx_sample": sorted(fx_serials)[:5],
                "ac_col": ac_map["serial"], "ac_sample": sorted(ac_serials)[:5],
            }
        }
    }
    state["last_reconcile"] = result
    return jsonify(result)

@app.route("/scan", methods=["POST"])
def scan():
    serial = request.json.get("serial", "").strip().upper()
    lookup = state.get("serial_lookup", {})
    code = lookup.get(serial)
    return jsonify({"serial": serial, "found": code is not None, "code": code or ""})

if __name__ == "__main__":
    app.run(debug=True, port=5050)
